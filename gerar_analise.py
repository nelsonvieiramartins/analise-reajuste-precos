import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ============================================================
# DADOS EXTRAIDOS DOS ARQUIVOS XLSX
# ============================================================

despesas_2025_jan = {
    'Despesas com funcionarias': 5440.59,
    'Porcentagem': 4812.76,
    'Aluguel': 3050.68,
    'Taxas e Despesas Diversas': 1932.35,
    'Emprestimo Juridico': 1854.00,
    'Materiais Odontologicos': 1407.13,
    'Laboratorios': 1300.00,
    'Contabilidade': 1214.40,
    'Juros dos Cartoes': 800.27,
    'Imposto Juridico': 583.49,
    'SSI - SESI': 286.51,
    'Telefone': 182.09,
    'ISS': 168.79,
    'Internet': 168.00,
    'Codental': 160.90,
    'CHB Ambiental': 95.65,
    'Agua': 88.96,
    'Material de conservacao e limpeza': 37.15,
}

despesas_2025_fev = {
    'Porcentagem': 6694.24,
    'Despesas com funcionarias': 5616.92,
    'Aluguel': 3006.00,
    'Laboratorios': 2940.00,
    'Juros dos Cartoes': 2823.61,
    'Materiais Odontologicos': 2007.63,
    'Emprestimo Juridico': 1854.00,
    'Energia eletrica': 1237.84,
    'Contabilidade': 1214.00,
    'Imposto Juridico': 708.05,
    'Taxas e Despesas Diversas': 420.18,
    'Telefone': 211.74,
    'Infraestrutura': 194.00,
    'ISS': 168.79,
    'Internet': 139.28,
    'Codental': 114.90,
    'Material de conservacao e limpeza': 96.30,
    'CHB Ambiental': 93.75,
    'Escritorio e Papelaria': 46.48,
}

despesas_2026_jan = {
    'Porcentagem': 4427.00,
    'Despesas com funcionarias': 4050.16,
    'Aluguel': 3007.72,
    'Emprestimo Juridico': 1766.00,
    'Materiais Odontologicos': 1715.56,
    'Laboratorios': 1500.00,
    'Energia eletrica': 1343.16,
    'Imposto Juridico': 1249.95,
    'Contabilidade': 1214.40,
    'Juros dos Cartoes': 825.60,
    'Publicidade/Propaganda': 650.00,
    'Manutencao de Maquinas e Equipamentos': 380.00,
    'Telefone': 296.45,
    'Material de conservacao e limpeza': 239.58,
    'CRO': 238.67,
    'Imposto Fiscal': 189.94,
    'Agua': 135.68,
    'Codental': 128.90,
    'Infraestrutura': 120.00,
    'CHB Ambiental': 101.17,
    'Marketing - Venda': 100.00,
    'NF Prefeitura': 40.87,
    'Escritorio e Papelaria': 40.00,
}

despesas_2026_fev = {
    'Despesas com funcionarias': 5869.27,
    'Porcentagem': 5242.00,
    'Laboratorios': 3150.00,
    'Aluguel': 2969.58,
    'Materiais Odontologicos': 2737.48,
    'Emprestimo Juridico': 1766.00,
    'Contabilidade': 1296.86,
    'Juros dos Cartoes': 1210.51,
    'Imposto Juridico': 1119.07,
    'Energia eletrica': 1032.64,
    'Manutencao de Maquinas e Equipamentos': 390.00,
    'Internet': 296.75,
    'CRO': 242.92,
    'Taxas e Despesas Diversas': 238.67,
    'ISS': 191.04,
    'Agua': 136.51,
    'Codental': 128.90,
    'CHB Ambiental': 101.27,
    'Material de conservacao e limpeza': 101.07,
    'NF Prefeitura': 79.80,
    'Telefone': 57.50,
}

# ============================================================
# DADOS DE RECEITA 2026 (Extraidos de RECEITA 2026/fluxo-de-caixa (3).xlsx)
# ============================================================
# Receita real mensal (faturamento com precos atuais)
receita_2026 = {
    'Janeiro': 26396.13,
    'Fevereiro': 36462.71,
    'Marco': 24411.66,
}

# Qtd de procedimentos realizados por mes
procedimentos_2026 = {
    'Janeiro': 77,
    'Fevereiro': 105,
    'Marco': 71,
}

# Classificacao dos custos
FIXOS = ['Aluguel', 'Contabilidade', 'Emprestimo Juridico', 'Internet', 'Telefone',
         'Codental', 'CHB Ambiental', 'SSI - SESI', 'CRO', 'ISS', 'NF Prefeitura', 'Imposto Fiscal']

SEMI_VARIAVEIS = ['Despesas com funcionarias', 'Energia eletrica', 'Agua',
                  'Imposto Juridico', 'Infraestrutura', 'Material de conservacao e limpeza',
                  'Escritorio e Papelaria', 'Manutencao de Maquinas e Equipamentos',
                  'Publicidade/Propaganda', 'Marketing - Venda']

VARIAVEIS = ['Porcentagem', 'Materiais Odontologicos', 'Laboratorios',
             'Juros dos Cartoes', 'Taxas e Despesas Diversas']

# Tabela de tratamentos
tabela_tratamentos = [
    ('Extracoes', 'Exodontia Simples', 200.00),
    ('Extracoes', 'Exodontia Siso Superiores', 275.00),
    ('Extracoes', 'Exodontia Siso Inferior', 350.00),
    ('Protese', 'Coroa Artglass / Ceramero', 600.00),
    ('Protese', 'Coroa de Ceramica Pura', 1200.00),
    ('Protese', 'Pino de Fibra de Vidro', 100.00),
    ('Protese', 'Adesiva (2 rest. + 1 suspensa)', 1100.00),
    ('Protese', 'Fixa 2 ou 3 elementos', 1500.00),
    ('Protese', 'Restaurado de Resina (protetico)', 500.00),
    ('Protese', 'Inlay/Onlay de Porcelana', 1000.00),
    ('Protese', 'PT (Protese Total)', 1200.00),
    ('Protese', 'PPR', 1200.00),
    ('Aparelhos', 'Manutencao', 100.00),
    ('Aparelhos', 'Autoligado', 700.00),
    ('Aparelhos', 'Estetico', 600.00),
    ('Aparelhos', 'Metalico', 300.00),
    ('Aparelhos', 'Contencao Movel', 350.00),
    ('Aparelhos', 'Alinhador', 250.00),
    ('Implante', 'Cirurgia de Implante', 1200.00),
    ('Implante', 'PSI', 1300.00),
    ('Endodontia', 'Canal Molar', 950.00),
    ('Endodontia', 'Canal Pre-molar', 850.00),
    ('Endodontia', 'Canal Incisivo', 850.00),
    ('Clareamento', '1 Sessao', 350.00),
    ('Clareamento', 'Caseiro', 400.00),
    ('Clareamento', 'Completo (2 sessoes + caseiro)', 1000.00),
    ('Facetas/Lentes', 'Faceta em Resina', 275.00),
    ('Facetas/Lentes', 'Lente de Porcelana', 1200.00),
    ('Restauracoes', 'Restauracao Simples', 140.00),
    ('Restauracoes', 'Restauracao de Resina', 225.00),
    ('Outros', 'Profilaxia + Raspagem', 250.00),
]

# Procedimentos que dependem de laboratorio
USA_LAB = {'Coroa Artglass / Ceramero', 'Coroa de Ceramica Pura', 'Adesiva (2 rest. + 1 suspensa)',
           'Fixa 2 ou 3 elementos', 'Restaurado de Resina (protetico)', 'Inlay/Onlay de Porcelana',
           'PT (Protese Total)', 'PPR', 'Lente de Porcelana'}

# ============================================================
# HELPERS
# ============================================================

BRL = 'R$ #,##0.00'
PCT = '0.0%'

def s(ws, r, c, val=None, bold=False, bg=None, fc="000000", al="left",
      fmt=None, sz=10, bdr=False, wrap=False):
    cell = ws.cell(row=r, column=c)
    if val is not None:
        cell.value = val
    cell.font = Font(bold=bold, color=fc, size=sz, name="Calibri")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=al, vertical="center", wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    if bdr:
        t = Side(style='thin', color='BFBFBF')
        cell.border = Border(left=t, right=t, top=t, bottom=t)
    return cell

def sum_cat(d, cats):
    return sum(d.get(c, 0) for c in cats)

def get_tipo(cat):
    if cat in FIXOS: return "Fixo"
    if cat in SEMI_VARIAVEIS: return "Semi-Var."
    if cat in VARIAVEIS: return "Variavel"
    return "Outros"

# Cores
C_TITULO  = "1F4E79"
C_HEADER  = "2E75B6"
C_FIXO    = "E2EFDA"
C_SEMI    = "FFF2CC"
C_VAR     = "FCE4D6"
C_VERDE   = "00B050"
C_VERM    = "C00000"
C_BRANCO  = "FFFFFF"

# Totais gerais
T25J = sum(despesas_2025_jan.values())
T25F = sum(despesas_2025_fev.values())
T26J = sum(despesas_2026_jan.values())
T26F = sum(despesas_2026_fev.values())
TOTAL_25 = T25J + T25F
TOTAL_26 = T26J + T26F

VAR_JAN = T26J / T25J - 1
VAR_FEV = T26F / T25F - 1
VAR_MEDIA = (VAR_JAN + VAR_FEV) / 2

LAB_25 = despesas_2025_jan.get('Laboratorios',0) + despesas_2025_fev.get('Laboratorios',0)
LAB_26 = despesas_2026_jan.get('Laboratorios',0) + despesas_2026_fev.get('Laboratorios',0)
VAR_LAB = LAB_26 / LAB_25 - 1

MAT_25 = despesas_2025_jan.get('Materiais Odontologicos',0) + despesas_2025_fev.get('Materiais Odontologicos',0)
MAT_26 = despesas_2026_jan.get('Materiais Odontologicos',0) + despesas_2026_fev.get('Materiais Odontologicos',0)
VAR_MAT = MAT_26 / MAT_25 - 1

REAJ_GERAL = 0.10   # arredondado 10%
REAJ_LAB   = 0.15   # laboratorios subiram mais

MED_FIXOS = (sum_cat(despesas_2026_jan, FIXOS) + sum_cat(despesas_2026_fev, FIXOS)) / 2
MED_SEMI  = (sum_cat(despesas_2026_jan, SEMI_VARIAVEIS) + sum_cat(despesas_2026_fev, SEMI_VARIAVEIS)) / 2
MED_VAR   = (sum_cat(despesas_2026_jan, VARIAVEIS) + sum_cat(despesas_2026_fev, VARIAVEIS)) / 2
MED_TOTAL = MED_FIXOS + MED_SEMI + MED_VAR

# ============================================================
# CRIAR WORKBOOK
# ============================================================
wb = openpyxl.Workbook()

# ============================================================
# ABA 1 - RESUMO GERAL
# ============================================================
ws1 = wb.active
ws1.title = "1. Resumo Geral"

# larguras
for i, w in enumerate([32, 13, 13, 15, 13, 13, 15, 14, 11], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# titulo
ws1.merge_cells("A1:I1")
s(ws1,1,1,"ANALISE DE DESPESAS - CLINICA ODONTOLOGICA PELEGRINO",
  bold=True, bg=C_TITULO, fc="FFFFFF", al="center", sz=13)
ws1.row_dimensions[1].height = 28

ws1.merge_cells("A2:I2")
s(ws1,2,1,"Comparativo Jan-Fev 2025 x Jan-Fev 2026  |  Base para Reajuste de Tabela de Precos",
  bg="BDD7EE", fc=C_TITULO, al="center", sz=10)
ws1.row_dimensions[2].height = 18

# cabecalho
r = 4
hdrs = ["CATEGORIA DE DESPESA","Tipo","Jan/25","Fev/25","TOTAL 2025","Jan/26","Fev/26","TOTAL 2026","Var. %"]
for c, h in enumerate(hdrs, 1):
    s(ws1, r, c, h, bold=True, bg=C_HEADER, fc="FFFFFF", al="center", bdr=True)
ws1.row_dimensions[r].height = 20

# todas as categorias ordenadas por tipo
todas = sorted(set(
    list(despesas_2025_jan) + list(despesas_2025_fev) +
    list(despesas_2026_jan) + list(despesas_2026_fev)
), key=lambda c: (get_tipo(c), c))

for cat in todas:
    r += 1
    j25 = despesas_2025_jan.get(cat, 0)
    f25 = despesas_2025_fev.get(cat, 0)
    t25 = j25 + f25
    j26 = despesas_2026_jan.get(cat, 0)
    f26 = despesas_2026_fev.get(cat, 0)
    t26 = j26 + f26
    vp  = (t26 / t25 - 1) if t25 > 0 else None

    tipo = get_tipo(cat)
    bg = C_FIXO if tipo == "Fixo" else (C_SEMI if tipo == "Semi-Var." else (C_VAR if tipo == "Variavel" else "F2F2F2"))

    s(ws1, r, 1, cat, bg=bg, bdr=True)
    s(ws1, r, 2, tipo, bg=bg, al="center", bdr=True, sz=9)
    for c, (v, fmt) in enumerate([(j25,BRL),(f25,BRL),(t25,BRL),(j26,BRL),(f26,BRL),(t26,BRL),(vp,PCT)], 3):
        cell = s(ws1, r, c, v if v else None, bg=bg, al="right", fmt=fmt, bdr=True)
        if c == 9 and vp is not None and vp > 0.08:
            cell.font = Font(color=C_VERM, size=10, bold=True, name="Calibri")
        elif c == 9 and vp is not None and vp < 0:
            cell.font = Font(color=C_VERDE, size=10, name="Calibri")

# subtotais
grupos = [
    ("SUBTOTAL CUSTOS FIXOS", FIXOS, "C6EFCE"),
    ("SUBTOTAL CUSTOS SEMI-VARIAVEIS", SEMI_VARIAVEIS, "FFEB9C"),
    ("SUBTOTAL CUSTOS VARIAVEIS", VARIAVEIS, "FFCC99"),
]
for label, cats, bg in grupos:
    r += 1
    j25 = sum_cat(despesas_2025_jan, cats)
    f25 = sum_cat(despesas_2025_fev, cats)
    t25 = j25 + f25
    j26 = sum_cat(despesas_2026_jan, cats)
    f26 = sum_cat(despesas_2026_fev, cats)
    t26 = j26 + f26
    vp  = t26 / t25 - 1
    for c, (v, fmt) in enumerate([(label,None),(None,None),(j25,BRL),(f25,BRL),(t25,BRL),(j26,BRL),(f26,BRL),(t26,BRL),(vp,PCT)], 1):
        s(ws1, r, c, v, bold=True, bg=bg, al="right" if c > 2 else "left", fmt=fmt, bdr=True)

# total geral
r += 1
vp_total = TOTAL_26 / TOTAL_25 - 1
for c, (v, fmt) in enumerate([("TOTAL GERAL",None),(None,None),(T25J,BRL),(T25F,BRL),(TOTAL_25,BRL),(T26J,BRL),(T26F,BRL),(TOTAL_26,BRL),(vp_total,PCT)], 1):
    s(ws1, r, c, v, bold=True, bg=C_TITULO, fc="FFFFFF", al="right" if c > 2 else "left", fmt=fmt, bdr=True, sz=11)
ws1.row_dimensions[r].height = 22

# legenda
r += 2
for bg, texto in [
    (C_FIXO,  "  FIXO: Despesas mensais certas — existem independente dos atendimentos (Aluguel, Contabilidade...)"),
    (C_SEMI,  "  SEMI-VARIAVEL: Mensais com variacao (Funcionarias, Energia, Agua, Impostos...)"),
    (C_VAR,   "  VARIAVEL: Ligados a producao — quanto mais se produz, mais se gasta (Comissoes, Materiais, Lab.)"),
]:
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    s(ws1, r, 1, texto, bg=bg, sz=9)
    r += 1

ws1.freeze_panes = "A5"

# ============================================================
# ABA 2 - CUSTOS FIXOS
# ============================================================
ws2 = wb.create_sheet("2. Custos Fixos")

for i, w in enumerate([32, 14, 14, 16, 14, 14, 16, 14, 11], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:I1")
s(ws2,1,1,"CUSTOS FIXOS E SEMI-VARIAVEIS MENSAIS",bold=True,bg=C_TITULO,fc="FFFFFF",al="center",sz=13)
ws2.row_dimensions[1].height = 28

ws2.merge_cells("A2:I2")
s(ws2,2,1,"Despesas recorrentes que precisam ser cobertas todo mes, independente do volume de atendimentos",
  bg="BDD7EE",fc=C_TITULO,al="center",sz=10)
ws2.row_dimensions[2].height = 18

r = 4
for c, h in enumerate(["DESPESA","Tipo","Jan/25","Fev/25","Media/25","Jan/26","Fev/26","Media/26","Var.%"], 1):
    s(ws2, r, c, h, bold=True, bg=C_HEADER, fc="FFFFFF", al="center", bdr=True)
ws2.row_dimensions[r].height = 20

for cat in sorted(FIXOS + SEMI_VARIAVEIS):
    r += 1
    j25 = despesas_2025_jan.get(cat,0)
    f25 = despesas_2025_fev.get(cat,0)
    m25 = (j25+f25)/2
    j26 = despesas_2026_jan.get(cat,0)
    f26 = despesas_2026_fev.get(cat,0)
    m26 = (j26+f26)/2
    vp  = (m26/m25-1) if m25 > 0 else None
    bg  = C_FIXO if cat in FIXOS else C_SEMI
    tipo = "Fixo" if cat in FIXOS else "Semi-Var."
    s(ws2,r,1,cat,bg=bg,bdr=True)
    s(ws2,r,2,tipo,bg=bg,al="center",bdr=True,sz=9)
    for c, (v,fmt) in enumerate([(j25,BRL),(f25,BRL),(m25,BRL),(j26,BRL),(f26,BRL),(m26,BRL),(vp,PCT)],3):
        cell = s(ws2,r,c,v if v else None,bg=bg,al="right",fmt=fmt,bdr=True)
        if c==9 and vp and vp > 0.05:
            cell.font = Font(color=C_VERM,size=10,name="Calibri")
        elif c==9 and vp and vp < 0:
            cell.font = Font(color=C_VERDE,size=10,name="Calibri")

# totais fixos+semi
r += 1
j25_fs = sum_cat(despesas_2025_jan, FIXOS+SEMI_VARIAVEIS)
f25_fs = sum_cat(despesas_2025_fev, FIXOS+SEMI_VARIAVEIS)
m25_fs = (j25_fs+f25_fs)/2
j26_fs = sum_cat(despesas_2026_jan, FIXOS+SEMI_VARIAVEIS)
f26_fs = sum_cat(despesas_2026_fev, FIXOS+SEMI_VARIAVEIS)
m26_fs = (j26_fs+f26_fs)/2
vp_fs  = m26_fs/m25_fs-1
s(ws2,r,1,"TOTAL FIXO + SEMI-VARIAVEL",bold=True,bg=C_TITULO,fc="FFFFFF",bdr=True,sz=11)
s(ws2,r,2,"",bold=True,bg=C_TITULO,fc="FFFFFF",bdr=True)
for c,(v,fmt) in enumerate([(j25_fs,BRL),(f25_fs,BRL),(m25_fs,BRL),(j26_fs,BRL),(f26_fs,BRL),(m26_fs,BRL),(vp_fs,PCT)],3):
    s(ws2,r,c,v,bold=True,bg=C_TITULO,fc="FFFFFF",al="right",fmt=fmt,bdr=True,sz=11)
ws2.row_dimensions[r].height = 22

r += 2
infos = [
    ("  PONTOS IMPORTANTES PARA PRECIFICACAO:", True, C_TITULO, "FFFFFF"),
    (f"  Custo fixo+semi medio mensal 2025: R$ {m25_fs:,.2f} — Este era o overhead base do ano passado.", False, "DEEAF1", "000000"),
    (f"  Custo fixo+semi medio mensal 2026: R$ {m26_fs:,.2f} — Variacao de {vp_fs:.1%} em relacao a 2025.", False, "DEEAF1", "000000"),
    (f"  Para cobrir apenas os custos fixos/semi em 2026, a clinica precisa faturar no minimo R$ {m26_fs:,.2f}/mes.", False, "FFF2CC", "000000"),
    ("  Os custos de materiais e laboratorio devem ser calculados individualmente por procedimento.", False, "FFF2CC", "000000"),
    ("  ATENCAO: Esses valores representam o overhead — nao incluem o lucro da clinica.", False, "FCE4D6", "000000"),
]
for txt, bold, bg, fc in infos:
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    s(ws2,r,1,txt,bold=bold,bg=bg,fc=fc,sz=9 if not bold else 10)
    ws2.row_dimensions[r].height = 18
    r += 1

ws2.freeze_panes = "A5"

# ============================================================
# ABA 3 - VARIACAO DETALHADA
# ============================================================
ws3 = wb.create_sheet("3. Variacao Detalhada")

for i, w in enumerate([32, 12, 13, 13, 13, 11, 13, 13, 13, 11], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:J1")
s(ws3,1,1,"VARIACAO DETALHADA POR CATEGORIA — MES A MES",bold=True,bg=C_TITULO,fc="FFFFFF",al="center",sz=13)
ws3.row_dimensions[1].height = 28
ws3.merge_cells("A2:J2")
s(ws3,2,1,"Compara cada despesa: Janeiro 2025 vs 2026 e Fevereiro 2025 vs 2026",bg="BDD7EE",fc=C_TITULO,al="center",sz=10)

r = 4
for c,h in enumerate(["CATEGORIA","Tipo","Jan/25","Jan/26","Var.Jan R$","Var.Jan%","Fev/25","Fev/26","Var.Fev R$","Var.Fev%"],1):
    s(ws3,r,c,h,bold=True,bg=C_HEADER,fc="FFFFFF",al="center",bdr=True)
ws3.row_dimensions[r].height = 20

todas2 = sorted(set(
    list(despesas_2025_jan)+list(despesas_2025_fev)+
    list(despesas_2026_jan)+list(despesas_2026_fev)
), key=lambda c:(get_tipo(c),c))

for cat in todas2:
    r += 1
    j25 = despesas_2025_jan.get(cat,0)
    j26 = despesas_2026_jan.get(cat,0)
    f25 = despesas_2025_fev.get(cat,0)
    f26 = despesas_2026_fev.get(cat,0)
    vj_r = j26-j25
    vj_p = (j26/j25-1) if j25>0 else None
    vf_r = f26-f25
    vf_p = (f26/f25-1) if f25>0 else None
    tipo = get_tipo(cat)
    bg = C_FIXO if tipo=="Fixo" else (C_SEMI if tipo=="Semi-Var." else (C_VAR if tipo=="Variavel" else "F2F2F2"))

    s(ws3,r,1,cat,bg=bg,bdr=True)
    s(ws3,r,2,tipo,bg=bg,al="center",bdr=True,sz=9)
    for c,(v,fmt) in enumerate([(j25,BRL),(j26,BRL),(vj_r if (j25>0 or j26>0) else None,BRL),(vj_p,PCT),(f25,BRL),(f26,BRL),(vf_r if (f25>0 or f26>0) else None,BRL),(vf_p,PCT)],3):
        cell = s(ws3,r,c,v if v else None,bg=bg,al="right",fmt=fmt,bdr=True)
        if c in [5,9] and v and v > 0:
            cell.font = Font(color=C_VERM,size=10,name="Calibri")
        elif c in [5,9] and v and v < 0:
            cell.font = Font(color=C_VERDE,size=10,name="Calibri")
        if c in [6,10] and v and v > 0.10:
            cell.font = Font(color=C_VERM,bold=True,size=10,name="Calibri")

r += 1
vj_total = T26J-T25J; vjp = T26J/T25J-1
vf_total = T26F-T25F; vfp = T26F/T25F-1
s(ws3,r,1,"TOTAL GERAL",bold=True,bg=C_TITULO,fc="FFFFFF",bdr=True)
s(ws3,r,2,"",bold=True,bg=C_TITULO,fc="FFFFFF",bdr=True)
for c,(v,fmt) in enumerate([(T25J,BRL),(T26J,BRL),(vj_total,BRL),(vjp,PCT),(T25F,BRL),(T26F,BRL),(vf_total,BRL),(vfp,PCT)],3):
    s(ws3,r,c,v,bold=True,bg=C_TITULO,fc="FFFFFF",al="right",fmt=fmt,bdr=True,sz=11)
ws3.row_dimensions[r].height = 22
ws3.freeze_panes = "A5"

# ============================================================
# ABA 4 - ANALISE DE PRECIFICACAO
# ============================================================
ws4 = wb.create_sheet("4. Analise de Precificacao")

for i, w in enumerate([38, 18, 15, 40], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells("A1:D1")
s(ws4,1,1,"ANALISE PARA REVISAO DE PRECOS",bold=True,bg=C_TITULO,fc="FFFFFF",al="center",sz=13)
ws4.row_dimensions[1].height = 28
ws4.merge_cells("A2:D2")
s(ws4,2,1,"Distribuicao dos custos e impacto na tabela de tratamentos — Dados de 2026",bg="BDD7EE",fc=C_TITULO,al="center",sz=10)

r = 4
ws4.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
s(ws4,r,1,"  1. ESTRUTURA DE CUSTOS MENSAIS (Media Jan-Fev 2026)",bold=True,bg=C_HEADER,fc="FFFFFF",sz=11)
ws4.row_dimensions[r].height = 22

r += 1
for c,h in enumerate(["BLOCO DE CUSTO","Valor Medio/Mes","% do Total","Observacao"],1):
    s(ws4,r,c,h,bold=True,bg=C_HEADER,fc="FFFFFF",al="center",bdr=True)

blocos = [
    ("Custos Fixos (Aluguel, Contab., ISS...)", MED_FIXOS, MED_FIXOS/MED_TOTAL, "Invariavel — existem sem atendimentos", C_FIXO, "000000"),
    ("Custos Semi-Variaveis (Func., Energia...)", MED_SEMI, MED_SEMI/MED_TOTAL, "Variam pouco — principalmente funcionarias", C_SEMI, "000000"),
    ("Custos Variaveis (Comissao, Materiais, Lab.)", MED_VAR, MED_VAR/MED_TOTAL, "Proporcionais a producao", C_VAR, "000000"),
    ("TOTAL MEDIO MENSAL 2026", MED_TOTAL, 1.0, "Despesa total a cobrir mensalmente", C_TITULO, "FFFFFF"),
]
for nome, val, pct, obs, bg, fc in blocos:
    r += 1
    bold = nome.startswith("TOTAL")
    s(ws4,r,1,nome,bold=bold,bg=bg,fc=fc,bdr=True)
    s(ws4,r,2,val,bold=bold,bg=bg,fc=fc,al="right",fmt=BRL,bdr=True)
    s(ws4,r,3,pct,bold=bold,bg=bg,fc=fc,al="center",fmt=PCT,bdr=True)
    s(ws4,r,4,obs,bg=bg,fc=fc,bdr=True)
    ws4.row_dimensions[r].height = 18

r += 2
ws4.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
s(ws4,r,1,"  2. CUSTO DE OVERHEAD POR HORA E POR PROCEDIMENTO",bold=True,bg=C_HEADER,fc="FFFFFF",sz=11)
ws4.row_dimensions[r].height = 22

HORAS_MES = 22 * 8  # 176h
C_HORA_FIXO = (MED_FIXOS + MED_SEMI) / HORAS_MES
C_HORA_TOTAL = MED_TOTAL / HORAS_MES
TEMPO_PROC = 1.5  # horas por procedimento (estimativa)
OVERHEAD_PROC = C_HORA_FIXO * TEMPO_PROC

r += 1
params = [
    ("Dias trabalhados por mes (estimado)", "22 dias", "Ajuste conforme agenda real"),
    ("Horas de atendimento por dia", "8 horas", "Inclui todos os dentistas"),
    ("Total horas/mes", f"{HORAS_MES} horas/mes", "22 x 8h"),
    (f"Custo fixo+semi por hora trabalhada", f"R$ {C_HORA_FIXO:.2f}/hora", "Overhead de estrutura"),
    (f"Custo total (incluindo variaveis) por hora", f"R$ {C_HORA_TOTAL:.2f}/hora", "Referencia geral"),
    (f"Tempo medio estimado por procedimento", f"{TEMPO_PROC}h", "Considere ajustar por tipo"),
    (f"OVERHEAD MINIMO POR PROCEDIMENTO", f"R$ {OVERHEAD_PROC:.2f}", "Custo de estrutura que cada procedimento deve cobrir"),
]

for c,h in enumerate(["PARAMETRO","VALOR ESTIMADO","OBSERVACAO"],1):
    s(ws4,r,c,h,bold=True,bg=C_HEADER,fc="FFFFFF",al="center",bdr=True)
r += 1

for nome, val, obs in params:
    bold = "OVERHEAD MINIMO" in nome
    bg = "C6EFCE" if bold else "F2F2F2"
    s(ws4,r,1,nome,bold=bold,bg=bg,bdr=True)
    s(ws4,r,2,val,bold=bold,bg=bg,al="center",bdr=True)
    s(ws4,r,3,obs,bg=bg,bdr=True,sz=9)
    ws4.row_dimensions[r].height = 18
    r += 1

r += 1
ws4.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
s(ws4,r,1,"  3. VARIACOES REAIS IDENTIFICADAS (Base para Reajuste)",bold=True,bg=C_HEADER,fc="FFFFFF",sz=11)
ws4.row_dimensions[r].height = 22

r += 1
for c,h in enumerate(["INDICADOR","VARIACAO MEDIDA","IMPACTO"],1):
    s(ws4,r,c,h,bold=True,bg=C_HEADER,fc="FFFFFF",al="center",bdr=True)
r += 1

indicadores = [
    ("Variacao total Jan/25 vs Jan/26", f"{VAR_JAN:.1%}", "Referencia para reajuste geral"),
    ("Variacao total Fev/25 vs Fev/26", f"{VAR_FEV:.1%}", "Referencia para reajuste geral"),
    ("Variacao media bimestral", f"{VAR_MEDIA:.1%}", "Base principal do calculo"),
    ("Variacao especifica: Laboratorios", f"{VAR_LAB:.1%}", "Impacta proteses, coroas, lentes"),
    ("Variacao especifica: Materiais Odontologicos", f"{VAR_MAT:.1%}", "Impacta restauracoes, extracao, endo"),
    ("Porcentagem dentistas (media mensal 26)", f"R$ {(despesas_2026_jan.get('Porcentagem',0)+despesas_2026_fev.get('Porcentagem',0))/2:,.2f}", "Custo variavel de comissoes — proporcional ao faturamento"),
    ("REAJUSTE GERAL SUGERIDO", f"{REAJ_GERAL:.0%}", "Arredondado acima da variacao real para cobrir inflacao"),
    ("REAJUSTE PARA PROTESES (usa laboratorio)", f"{REAJ_LAB:.0%}", "Reflete variacao maior dos laboratorios"),
]

for nome, val, obs in indicadores:
    bold = "REAJUSTE" in nome
    bg = "C6EFCE" if bold else "DEEAF1"
    s(ws4,r,1,nome,bold=bold,bg=bg,bdr=True)
    s(ws4,r,2,val,bold=bold,bg=bg,al="center",bdr=True)
    s(ws4,r,3,obs,bg=bg,bdr=True,sz=9)
    ws4.row_dimensions[r].height = 18
    r += 1

r += 1
ws4.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
s(ws4,r,1,"  IMPORTANTE: Para refinar a analise, informe o numero medio de procedimentos por mes e o faturamento medio mensal.",
  bg="FCE4D6",fc="7F0000",sz=9,wrap=True)
ws4.row_dimensions[r].height = 24

ws4.freeze_panes = "A4"

# ============================================================
# ABA 5 - TABELA DE PRECOS
# ============================================================
ws5 = wb.create_sheet("5. Tabela de Precos")

for i, w in enumerate([16, 38, 14, 12, 14, 14, 12, 42], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells("A1:H1")
s(ws5,1,1,"TABELA DE PRECOS — ATUAL E PROPOSTA DE REAJUSTE",bold=True,bg=C_TITULO,fc="FFFFFF",al="center",sz=13)
ws5.row_dimensions[1].height = 28

ws5.merge_cells("A2:H2")
s(ws5,2,1,f"Reajuste calculado com base na variacao real das despesas Jan-Fev 2025 x 2026 | Reajuste geral: {REAJ_GERAL:.0%} | Proteses/Lab: {REAJ_LAB:.0%}",
  bg="BDD7EE",fc=C_TITULO,al="center",sz=10)
ws5.row_dimensions[2].height = 18

r = 4
for c,h in enumerate(["GRUPO","PROCEDIMENTO","PRECO ATUAL","REAJUSTE","PRECO CALCULADO","PRECO SUGERIDO","DIFERENCA","OBSERVACAO"],1):
    s(ws5,r,c,h,bold=True,bg=C_HEADER,fc="FFFFFF",al="center",bdr=True)
ws5.row_dimensions[r].height = 20

cores_grupos = {
    'Extracoes': "DEEAF1", 'Protese': "FCE4D6", 'Aparelhos': "E2EFDA",
    'Implante': "FFF2CC", 'Endodontia': "F2CEEF", 'Clareamento': "DDEEFF",
    'Facetas/Lentes': "FCE4D6", 'Restauracoes': "E2EFDA", 'Outros': "F5F5F5"
}

grupo_atual = ""
for grupo, proc, preco in tabela_tratamentos:
    r += 1
    reajuste = REAJ_LAB if proc in USA_LAB else REAJ_GERAL
    preco_calc = preco * (1 + reajuste)
    # arredondamento bonito
    if preco_calc < 150:    arred = round(preco_calc / 10) * 10
    elif preco_calc < 300:  arred = round(preco_calc / 25) * 25
    elif preco_calc < 600:  arred = round(preco_calc / 50) * 50
    else:                   arred = round(preco_calc / 100) * 100
    diff = arred - preco

    if proc in USA_LAB:
        obs = f"Usa laboratorio — reaj. {reajuste:.0%} (lab. subiu {VAR_LAB:.1%})"
    elif grupo == 'Implante':
        obs = "Verificar custo do componente + laboratorio separadamente"
    elif proc == 'Manutencao':
        obs = "Reajuste menor indicado — procedimento mensal, impacta fidelizacao"
    else:
        obs = f"Reaj. {reajuste:.0%} s/ variacao real das despesas ({VAR_MEDIA:.1%})"

    bg = cores_grupos.get(grupo, "FFFFFF")
    mostrar_grupo = grupo if grupo != grupo_atual else ""
    grupo_atual = grupo

    s(ws5,r,1,mostrar_grupo,bold=(mostrar_grupo!=""),bg=bg,bdr=True)
    s(ws5,r,2,proc,bg=bg,bdr=True)
    s(ws5,r,3,preco,bg=bg,al="right",fmt=BRL,bdr=True)
    s(ws5,r,4,reajuste,bg=bg,al="center",fmt=PCT,bdr=True)
    s(ws5,r,5,preco_calc,bg=bg,al="right",fmt=BRL,bdr=True)
    s(ws5,r,6,arred,bold=True,bg=bg,al="right",fmt=BRL,bdr=True)
    s(ws5,r,7,diff,bg=bg,al="right",fmt=BRL,bdr=True)
    s(ws5,r,8,obs,bg=bg,sz=9,bdr=True,wrap=True)
    ws5.row_dimensions[r].height = 18

r += 2
ws5.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
s(ws5,r,1,"  METODOLOGIA E NOTAS:",bold=True,bg=C_TITULO,fc="FFFFFF")
ws5.row_dimensions[r].height = 20

notas = [
    f"  Reajuste geral de {REAJ_GERAL:.0%}: variacao real das despesas foi {VAR_MEDIA:.1%} (media jan+fev). Adicionado margem para cobrir inflacao do periodo.",
    f"  Reajuste proteses/laboratorio de {REAJ_LAB:.0%}: laboratorios variaram {VAR_LAB:.1%} entre os periodos — maior que a media.",
    f"  Materiais odontologicos variaram {VAR_MAT:.1%} — impacta restauracoes, extrações e endodontia.",
    "  Preco Sugerido ja esta arredondado para valores comercialmente praticos (multiplos de R$ 25, 50 ou 100).",
    "  Para ajuste fino: informe o numero de procedimentos por mes para calcular o overhead individual.",
    "  Recomendacao: Revisar trimestralmente com base no fluxo de caixa atualizado.",
]
for nota in notas:
    ws5.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
    s(ws5,r,1,nota,bg="DEEAF1",sz=9,wrap=True)
    ws5.row_dimensions[r].height = 16
    r += 1

ws5.freeze_panes = "A5"

# ============================================================
# ABA 6 - DASHBOARD
# ============================================================
ws6 = wb.create_sheet("6. Dashboard")

for i, w in enumerate([30, 16, 16, 16, 16, 3, 30, 20], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

ws6.merge_cells("A1:H1")
s(ws6,1,1,"DASHBOARD — RESUMO EXECUTIVO",bold=True,bg=C_TITULO,fc="FFFFFF",al="center",sz=14)
ws6.row_dimensions[1].height = 32

ws6.merge_cells("A2:H2")
s(ws6,2,1,"Clinica Odontologica Pelegrino  |  Analise Jan-Fev 2025 x Jan-Fev 2026",bg="BDD7EE",fc=C_TITULO,al="center",sz=10)
ws6.row_dimensions[2].height = 18

# Quadro resumo por mes
r = 4
ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
s(ws6,r,1,"  DESPESAS POR BLOCO DE CUSTO",bold=True,bg=C_HEADER,fc="FFFFFF",sz=11)
ws6.row_dimensions[r].height = 22

r += 1
for c,h in enumerate(["","Jan/25","Fev/25","Jan/26","Fev/26"],1):
    s(ws6,r,c,h,bold=True,bg=C_HEADER,fc="FFFFFF",al="center",bdr=True)

meses_resumo = [
    ("Custos Fixos", C_FIXO,
     sum_cat(despesas_2025_jan,FIXOS), sum_cat(despesas_2025_fev,FIXOS),
     sum_cat(despesas_2026_jan,FIXOS), sum_cat(despesas_2026_fev,FIXOS)),
    ("Custos Semi-Variaveis", C_SEMI,
     sum_cat(despesas_2025_jan,SEMI_VARIAVEIS), sum_cat(despesas_2025_fev,SEMI_VARIAVEIS),
     sum_cat(despesas_2026_jan,SEMI_VARIAVEIS), sum_cat(despesas_2026_fev,SEMI_VARIAVEIS)),
    ("Custos Variaveis", C_VAR,
     sum_cat(despesas_2025_jan,VARIAVEIS), sum_cat(despesas_2025_fev,VARIAVEIS),
     sum_cat(despesas_2026_jan,VARIAVEIS), sum_cat(despesas_2026_fev,VARIAVEIS)),
    ("TOTAL GERAL", C_TITULO, T25J, T25F, T26J, T26F),
]
for nome, bg, j25, f25, j26, f26 in meses_resumo:
    r += 1
    bold = nome == "TOTAL GERAL"
    fc = "FFFFFF" if bold else "000000"
    s(ws6,r,1,nome,bold=bold,bg=bg,fc=fc,bdr=True)
    for c, v in enumerate([j25,f25,j26,f26], 2):
        s(ws6,r,c,v,bold=bold,bg=bg,fc=fc,al="right",fmt=BRL,bdr=True)

# KPIs
r += 2
ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
s(ws6,r,1,"  INDICADORES CHAVE",bold=True,bg=C_HEADER,fc="FFFFFF",sz=11)
ws6.row_dimensions[r].height = 22

r += 1
for c,h in enumerate(["INDICADOR","VALOR","STATUS","NOTA"],1):
    s(ws6,r,c,h,bold=True,bg=C_HEADER,fc="FFFFFF",al="center",bdr=True)
ws6.merge_cells(start_row=r,start_column=4,end_row=r,end_column=8)

kpis = [
    ("Total Bimestral 2025 (Jan+Fev)", TOTAL_25, BRL, "", "Total de despesas Jan-Fev 2025"),
    ("Total Bimestral 2026 (Jan+Fev)", TOTAL_26, BRL, "", "Total de despesas Jan-Fev 2026"),
    ("Variacao Total 2025 vs 2026", TOTAL_26-TOTAL_25, BRL, "Reducao de gastos" if TOTAL_26<TOTAL_25 else "Aumento de gastos", ""),
    ("Variacao Percentual Geral", (TOTAL_26-TOTAL_25)/TOTAL_25, PCT, "", ""),
    ("Media Mensal 2025", TOTAL_25/2, BRL, "", "Jan+Fev dividido por 2"),
    ("Media Mensal 2026", TOTAL_26/2, BRL, "", "Jan+Fev dividido por 2"),
    ("Overhead fixo+semi (media mensal 26)", MED_FIXOS+MED_SEMI, BRL, "Valor minimo a faturar", ""),
    ("Overhead total (media mensal 26)", MED_TOTAL, BRL, "Inclui custos variaveis", ""),
    ("Variacao Laboratorios (bimestral)", VAR_LAB, PCT, "Alta variacao" if VAR_LAB > 0.10 else "Estavel", ""),
    ("Variacao Materiais Odontologicos", VAR_MAT, PCT, "", ""),
    ("REAJUSTE GERAL APLICADO", REAJ_GERAL, PCT, "Aplicar a todos os precos", "Exceto proteses"),
    ("REAJUSTE PROTESES/LABORATORIO", REAJ_LAB, PCT, "Aplicar a proteses e coroas", ""),
]

for nome, val, fmt, status, nota in kpis:
    r += 1
    bold = "REAJUSTE" in nome
    bg = "C6EFCE" if bold else ("FFF8E1" if "Variacao" in nome else "F9F9F9")
    s(ws6,r,1,nome,bold=bold,bg=bg,bdr=True)
    s(ws6,r,2,val,bold=bold,bg=bg,al="right",fmt=fmt,bdr=True)
    s(ws6,r,3,status,bg=bg,al="center",bdr=True,sz=9)
    ws6.merge_cells(start_row=r,start_column=4,end_row=r,end_column=8)
    s(ws6,r,4,nota,bg=bg,bdr=True,sz=9)

r += 2
ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
s(ws6,r,1,"  PROXIMOS PASSOS RECOMENDADOS:",bold=True,bg=C_TITULO,fc="FFFFFF")
ws6.row_dimensions[r].height = 22

proximos = [
    "  1. Aplique o reajuste sugerido da Aba '5. Tabela de Precos' — revise cada procedimento com base na realidade da clinica.",
    "  2. Levante o numero medio de procedimentos realizados por mes para calcular o overhead por atendimento.",
    "  3. Identifique quais procedimentos sao mais frequentes e priorize o reajuste neles.",
    "  4. Considere uma margem de lucro de 20-30% ACIMA do custo total para sustentabilidade financeira.",
    "  5. Solicite dados de faturamento para cruzar receita x despesa e calcular o ponto de equilibrio.",
]
for p in proximos:
    ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
    s(ws6,r,1,p,bg="DEEAF1",sz=9,wrap=True)
    ws6.row_dimensions[r].height = 18
    r += 1

ws6.freeze_panes = "A4"

# ============================================================
# ABA 7 - SIMULACAO DE FATURAMENTO COM REAJUSTE
# ============================================================
ws7 = wb.create_sheet(u"7. Simula\u00e7\u00e3o Faturamento")

for i, w in enumerate([34, 16, 16, 16, 16, 16, 14, 42], 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

ws7.merge_cells("A1:H1")
s(ws7,1,1,u"SIMULA\u00c7\u00c3O DE FATURAMENTO — SE O REAJUSTE FOSSE APLICADO DESDE JAN/2026",
  bold=True, bg=C_TITULO, fc="FFFFFF", al="center", sz=13)
ws7.row_dimensions[1].height = 32

ws7.merge_cells("A2:H2")
s(ws7,2,1,f"Projecao de receita com reajuste geral de {REAJ_GERAL:.0%} aplicado sobre o faturamento real dos 3 primeiros meses de 2026",
  bg="BDD7EE", fc=C_TITULO, al="center", sz=10)
ws7.row_dimensions[2].height = 20

# --- Secao 1: Resumo mensal ---
r7 = 4
ws7.merge_cells(start_row=r7, start_column=1, end_row=r7, end_column=8)
s(ws7,r7,1,u"  1. COMPARA\u00c7\u00c3O MENSAL: RECEITA REAL x RECEITA COM REAJUSTE",
  bold=True, bg=C_HEADER, fc="FFFFFF", sz=11)
ws7.row_dimensions[r7].height = 22

r7 += 1
for c, h in enumerate([u"M\u00caS", "QTD PROC.", "RECEITA REAL",
                        f"RECEITA C/ REAJUSTE ({REAJ_GERAL:.0%})",
                        u"DIFEREN\u00c7A R$", u"DIFEREN\u00c7A %",
                        "TICKET MEDIO REAL", u"OBSERVA\u00c7\u00c3O"], 1):
    s(ws7, r7, c, h, bold=True, bg=C_HEADER, fc="FFFFFF", al="center", bdr=True)
ws7.row_dimensions[r7].height = 20

total_real = 0
total_simulado = 0
total_procs = 0

meses_ordem = ['Janeiro', 'Fevereiro', 'Marco']
meses_label = {'Janeiro': 'Janeiro/2026', 'Fevereiro': 'Fevereiro/2026', 'Marco': u'Mar\u00e7o/2026'}

for mes in meses_ordem:
    r7 += 1
    real = receita_2026[mes]
    procs = procedimentos_2026[mes]
    simulado = real * (1 + REAJ_GERAL)
    diff_r = simulado - real
    diff_p = REAJ_GERAL
    ticket_real = real / procs if procs > 0 else 0

    total_real += real
    total_simulado += simulado
    total_procs += procs

    bg = "E2EFDA" if mes != 'Marco' else "DEEAF1"
    obs = f"{procs} procedimentos realizados"

    s(ws7,r7,1, meses_label[mes], bg=bg, bdr=True, bold=True)
    s(ws7,r7,2, procs, bg=bg, al="center", bdr=True)
    s(ws7,r7,3, real, bg=bg, al="right", fmt=BRL, bdr=True)
    s(ws7,r7,4, simulado, bg=bg, al="right", fmt=BRL, bdr=True, bold=True)
    s(ws7,r7,5, diff_r, bg=bg, al="right", fmt=BRL, bdr=True)
    cell = s(ws7,r7,6, diff_p, bg=bg, al="center", fmt=PCT, bdr=True)
    cell.font = Font(color=C_VERDE, bold=True, size=10, name="Calibri")
    s(ws7,r7,7, ticket_real, bg=bg, al="right", fmt=BRL, bdr=True)
    s(ws7,r7,8, obs, bg=bg, bdr=True, sz=9)
    ws7.row_dimensions[r7].height = 20

# Linha TOTAL
r7 += 1
total_diff = total_simulado - total_real
total_diff_p = total_simulado / total_real - 1 if total_real > 0 else 0
ticket_total = total_real / total_procs if total_procs > 0 else 0

s(ws7,r7,1, "TOTAL TRIMESTRE", bold=True, bg=C_TITULO, fc="FFFFFF", bdr=True, sz=11)
s(ws7,r7,2, total_procs, bold=True, bg=C_TITULO, fc="FFFFFF", al="center", bdr=True, sz=11)
s(ws7,r7,3, total_real, bold=True, bg=C_TITULO, fc="FFFFFF", al="right", fmt=BRL, bdr=True, sz=11)
s(ws7,r7,4, total_simulado, bold=True, bg=C_TITULO, fc="FFFFFF", al="right", fmt=BRL, bdr=True, sz=11)
s(ws7,r7,5, total_diff, bold=True, bg=C_TITULO, fc="FFFFFF", al="right", fmt=BRL, bdr=True, sz=11)
s(ws7,r7,6, total_diff_p, bold=True, bg=C_TITULO, fc="FFFFFF", al="center", fmt=PCT, bdr=True, sz=11)
s(ws7,r7,7, ticket_total, bold=True, bg=C_TITULO, fc="FFFFFF", al="right", fmt=BRL, bdr=True, sz=11)
s(ws7,r7,8, "Soma dos 3 primeiros meses", bold=True, bg=C_TITULO, fc="FFFFFF", bdr=True, sz=9)
ws7.row_dimensions[r7].height = 24

# --- Secao 2: Analise do impacto ---
r7 += 2
ws7.merge_cells(start_row=r7, start_column=1, end_row=r7, end_column=8)
s(ws7,r7,1,"  2. IMPACTO FINANCEIRO DO REAJUSTE NO TRIMESTRE",
  bold=True, bg=C_HEADER, fc="FFFFFF", sz=11)
ws7.row_dimensions[r7].height = 22

r7 += 1
for c, h in enumerate(["INDICADOR", "VALOR", u"OBSERVA\u00c7\u00c3O"], 1):
    s(ws7, r7, c, h, bold=True, bg=C_HEADER, fc="FFFFFF", al="center", bdr=True)
ws7.merge_cells(start_row=r7, start_column=3, end_row=r7, end_column=8)

media_real = total_real / 3
media_sim = total_simulado / 3
ganho_mensal = media_sim - media_real
ganho_anual_proj = ganho_mensal * 12

impacto_items = [
    ("Receita real total (Jan-Mar/2026)", total_real, BRL, "Faturamento efetivo com precos atuais"),
    ("Receita simulada total (Jan-Mar/2026)", total_simulado, BRL, f"Se o reajuste de {REAJ_GERAL:.0%} ja estivesse em vigor"),
    ("Receita perdida no trimestre", total_diff, BRL, f"Valor que a clinica deixou de faturar por nao ter reajustado"),
    (u"M\u00e9dia mensal REAL", media_real, BRL, "Faturamento medio mensal com precos atuais"),
    (u"M\u00e9dia mensal COM REAJUSTE", media_sim, BRL, f"Faturamento medio mensal com reajuste de {REAJ_GERAL:.0%}"),
    ("Ganho mensal estimado com reajuste", ganho_mensal, BRL, "Diferenca media por mes"),
    (u"PROJE\u00c7\u00c3O ANUAL do ganho adicional", ganho_anual_proj, BRL, f"Se o reajuste for mantido nos 12 meses = R$ {ganho_anual_proj:,.2f}"),
    ("Despesa media mensal 2026", MED_TOTAL, BRL, "Para comparacao com a receita"),
]

for nome, val, fmt, obs in impacto_items:
    r7 += 1
    bold = "PROJE" in nome or "perdida" in nome.lower()
    bg = "C6EFCE" if bold else ("FFF2CC" if "Despesa" in nome else "F2F2F2")
    fc_color = C_VERM if "perdida" in nome.lower() else "000000"
    s(ws7,r7,1, nome, bold=bold, bg=bg, bdr=True)
    cell = s(ws7,r7,2, val, bold=bold, bg=bg, al="right", fmt=fmt, bdr=True)
    if "perdida" in nome.lower():
        cell.font = Font(color=C_VERM, bold=True, size=10, name="Calibri")
    ws7.merge_cells(start_row=r7, start_column=3, end_row=r7, end_column=8)
    s(ws7,r7,3, obs, bg=bg, bdr=True, sz=9)
    ws7.row_dimensions[r7].height = 20

# --- Secao 3: Margem operacional ---
r7 += 2
ws7.merge_cells(start_row=r7, start_column=1, end_row=r7, end_column=8)
s(ws7,r7,1,"  3. COMPARATIVO: RECEITA x DESPESA (Media Mensal)",
  bold=True, bg=C_HEADER, fc="FFFFFF", sz=11)
ws7.row_dimensions[r7].height = 22

r7 += 1
for c, h in enumerate([u"CEN\u00c1RIO", "RECEITA MENSAL", "DESPESA MENSAL", "SALDO", "MARGEM %", "", u"SITUA\u00c7\u00c3O"], 1):
    s(ws7, r7, c, h, bold=True, bg=C_HEADER, fc="FFFFFF", al="center", bdr=True)
ws7.merge_cells(start_row=r7, start_column=7, end_row=r7, end_column=8)

# Cenario atual
r7 += 1
saldo_atual = media_real - MED_TOTAL
margem_atual = saldo_atual / media_real if media_real > 0 else 0
bg_atual = "FCE4D6" if saldo_atual < 0 else "E2EFDA"
fc_saldo = C_VERM if saldo_atual < 0 else C_VERDE
sit_atual = u"D\u00c9FICIT" if saldo_atual < 0 else u"SUPER\u00c1VIT"

s(ws7,r7,1, u"Sem reajuste (pre\u00e7os atuais)", bg=bg_atual, bdr=True, bold=True)
s(ws7,r7,2, media_real, bg=bg_atual, al="right", fmt=BRL, bdr=True)
s(ws7,r7,3, MED_TOTAL, bg=bg_atual, al="right", fmt=BRL, bdr=True)
cell = s(ws7,r7,4, saldo_atual, bg=bg_atual, al="right", fmt=BRL, bdr=True, bold=True)
cell.font = Font(color=fc_saldo, bold=True, size=10, name="Calibri")
s(ws7,r7,5, margem_atual, bg=bg_atual, al="center", fmt=PCT, bdr=True)
s(ws7,r7,6, "", bg=bg_atual, bdr=True)
ws7.merge_cells(start_row=r7, start_column=7, end_row=r7, end_column=8)
cell = s(ws7,r7,7, sit_atual, bg=bg_atual, al="center", bdr=True, bold=True)
cell.font = Font(color=fc_saldo, bold=True, size=11, name="Calibri")
ws7.row_dimensions[r7].height = 22

# Cenario com reajuste
r7 += 1
saldo_sim = media_sim - MED_TOTAL
margem_sim = saldo_sim / media_sim if media_sim > 0 else 0
bg_sim = "FCE4D6" if saldo_sim < 0 else "E2EFDA"
fc_saldo_sim = C_VERM if saldo_sim < 0 else C_VERDE
sit_sim = u"D\u00c9FICIT" if saldo_sim < 0 else u"SUPER\u00c1VIT"

s(ws7,r7,1, f"Com reajuste de {REAJ_GERAL:.0%}", bg=bg_sim, bdr=True, bold=True)
s(ws7,r7,2, media_sim, bg=bg_sim, al="right", fmt=BRL, bdr=True)
s(ws7,r7,3, MED_TOTAL, bg=bg_sim, al="right", fmt=BRL, bdr=True)
cell = s(ws7,r7,4, saldo_sim, bg=bg_sim, al="right", fmt=BRL, bdr=True, bold=True)
cell.font = Font(color=fc_saldo_sim, bold=True, size=10, name="Calibri")
s(ws7,r7,5, margem_sim, bg=bg_sim, al="center", fmt=PCT, bdr=True)
s(ws7,r7,6, "", bg=bg_sim, bdr=True)
ws7.merge_cells(start_row=r7, start_column=7, end_row=r7, end_column=8)
cell = s(ws7,r7,7, sit_sim, bg=bg_sim, al="center", bdr=True, bold=True)
cell.font = Font(color=fc_saldo_sim, bold=True, size=11, name="Calibri")
ws7.row_dimensions[r7].height = 22

# --- Notas finais ---
r7 += 2
ws7.merge_cells(start_row=r7, start_column=1, end_row=r7, end_column=8)
s(ws7,r7,1,u"  CONCLUS\u00d5ES E NOTAS:", bold=True, bg=C_TITULO, fc="FFFFFF")
ws7.row_dimensions[r7].height = 22

notas_sim = [
    f"  Os dados de receita foram extraidos do arquivo RECEITA 2026/fluxo-de-caixa (3).xlsx.",
    f"  A simulacao aplica o reajuste geral de {REAJ_GERAL:.0%} uniformemente sobre todo o faturamento.",
    f"  Na pratica, procedimentos com laboratorio teriam reajuste de {REAJ_LAB:.0%}, aumentando ainda mais o faturamento simulado.",
    f"  Receita perdida acumulada no trimestre por nao reajustar: R$ {total_diff:,.2f}.",
    f"  Se o reajuste fosse aplicado o ano inteiro, o ganho adicional estimado seria de R$ {ganho_anual_proj:,.2f}.",
    f"  RECOMENDACAO: Aplicar o reajuste o mais breve possivel para recuperar a margem operacional.",
]
for nota in notas_sim:
    ws7.merge_cells(start_row=r7, start_column=1, end_row=r7, end_column=8)
    bold_nota = "RECOMENDACAO" in nota
    bg_nota = "FFF2CC" if bold_nota else "DEEAF1"
    s(ws7,r7,1, nota, bg=bg_nota, sz=9, bold=bold_nota, wrap=True)
    ws7.row_dimensions[r7].height = 18
    r7 += 1

ws7.freeze_panes = "A6"

# ============================================================
# SALVAR
# ============================================================
output = r"c:\Users\nelso\OneDrive\Pelegrino\_ANALISE DE REAJUSTE DE PREÇOS\ANALISE_REAJUSTE_COMPLETA.xlsx"
wb.save(output)
print("SALVO COM SUCESSO!")
print(f"Arquivo: {output}")
print(f"\n--- NUMEROS FINAIS ---")
print(f"Total despesas 2025 (Jan+Fev): R$ {TOTAL_25:,.2f}")
print(f"Total despesas 2026 (Jan+Fev): R$ {TOTAL_26:,.2f}")
print(f"Variacao: {(TOTAL_26/TOTAL_25-1):.2%}")
print(f"Media mensal 2025: R$ {TOTAL_25/2:,.2f}")
print(f"Media mensal 2026: R$ {TOTAL_26/2:,.2f}")
print(f"Overhead fixo+semi mensal 2026: R$ {MED_FIXOS+MED_SEMI:,.2f}")
print(f"Variacao laboratorios: {VAR_LAB:.2%}")
print(f"Variacao materiais: {VAR_MAT:.2%}")
print(f"Reajuste geral sugerido: {REAJ_GERAL:.0%}")
print(f"Reajuste proteses sugerido: {REAJ_LAB:.0%}")
